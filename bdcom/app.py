import streamlit as st
st.set_page_config(page_title="FRY14M Field Analysis", layout="wide", initial_sidebar_state="expanded")

import pandas as pd
import os
import datetime
import re
from dateutil.relativedelta import relativedelta
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, DataReturnMode
from st_aggrid.shared import JsCode
from openpyxl import load_workbook
from openpyxl.comments import Comment

#############################################
# 1) Helper Functions
#############################################

def compute_grid_height(df, row_height=40, header_height=80):
    """Compute approximate AgGrid height."""
    n = len(df)
    return header_height + (min(n, 30) * row_height)

def drop_prev_comments_columns(df):
    """
    Drop any columns that contain 'm- Comments' 
    so we don't accumulate duplicates like '2025-12 m- Comments_x'.
    """
    pattern = re.compile(r"m- Comments", re.IGNORECASE)
    drop_cols = [c for c in df.columns if pattern.search(str(c))]
    if drop_cols:
        df = df.drop(columns=drop_cols)
    return df

def flatten_dataframe(df):
    """Flatten multi-index columns into single-level columns."""
    if isinstance(df.columns, pd.MultiIndex):
        df = df.reset_index()
        df.columns = [
            " ".join(map(str, col)).strip() if isinstance(col, tuple) else col
            for col in df.columns.values
        ]
    return df

def normalize_columns(df, mapping={"field_name": "Field Name", "value_label": "Value Label"}):
    """
    Convert e.g. 'field_name' -> 'Field Name', 'value_label' -> 'Value Label', etc.
    Also strip whitespace from column names.
    """
    df.columns = [str(col).strip() for col in df.columns]
    for orig, new in mapping.items():
        for col in df.columns:
            if col.lower() == orig.lower() and col != new:
                df.rename(columns={col: new}, inplace=True)
    return df

#############################################
# 2) Summaries & Distribution
#############################################

def generate_summary_df(df_data, date1, date2):
    """Create the Summary sheet from the Data sheet (lowercase columns)."""
    fields = sorted(df_data["field_name"].unique())
    rows = []
    for field in fields:
        # Missing checks
        mask_miss_d1 = ((df_data['analysis_type'] == 'value_dist') &
                        (df_data['field_name'] == field) &
                        (df_data['filemonth_dt'] == date1) &
                        (df_data['value_label'].str.contains("Missing", case=False, na=False)))
        missing_d1 = df_data.loc[mask_miss_d1, 'value_records'].sum()

        mask_miss_d2 = ((df_data['analysis_type'] == 'value_dist') &
                        (df_data['field_name'] == field) &
                        (df_data['filemonth_dt'] == date2) &
                        (df_data['value_label'].str.contains("Missing", case=False, na=False)))
        missing_d2 = df_data.loc[mask_miss_d2, 'value_records'].sum()

        # Phrases for pop_comp
        phrases = [
            "1\\)   CF Loan - Both Pop, Diff Values",
            "2\\)   CF Loan - Prior Null, Current Pop",
            "3\\)   CF Loan - Prior Pop, Current Null"
        ]
        def contains_phrase(x):
            for pat in phrases:
                if pd.notna(x) and re.search(pat, x):
                    return True
            return False

        mask_pop_d1 = ((df_data['analysis_type'] == 'pop_comp') &
                       (df_data['field_name'] == field) &
                       (df_data['filemonth_dt'] == date1) &
                       (df_data['value_label'].apply(contains_phrase)))
        pop_d1 = df_data.loc[mask_pop_d1, 'value_records'].sum()

        mask_pop_d2 = ((df_data['analysis_type'] == 'pop_comp') &
                       (df_data['field_name'] == field) &
                       (df_data['filemonth_dt'] == date2) &
                       (df_data['value_label'].apply(contains_phrase)))
        pop_d2 = df_data.loc[mask_pop_d2, 'value_records'].sum()

        rows.append([field, missing_d1, missing_d2, pop_d1, pop_d2])

    summary = pd.DataFrame(rows, columns=[
        "Field Name",
        f"Missing Values ({date1.strftime('%Y-%m-%d')})",
        f"Missing Values ({date2.strftime('%Y-%m-%d')})",
        f"Month-to-Month Diff ({date1.strftime('%Y-%m-%d')})",
        f"Month-to-Month Diff ({date2.strftime('%Y-%m-%d')})"
    ])

    m1 = f"Missing Values ({date1.strftime('%Y-%m-%d')})"
    m2 = f"Missing Values ({date2.strftime('%Y-%m-%d')})"
    d1 = f"Month-to-Month Diff ({date1.strftime('%Y-%m-%d')})"
    d2 = f"Month-to-Month Diff ({date2.strftime('%Y-%m-%d')})"

    summary["Missing % Change"] = summary.apply(lambda r: ((r[m1]-r[m2]) / r[m2] * 100) if r[m2]!=0 else None, axis=1)
    summary["Month-to-Month % Change"] = summary.apply(lambda r: ((r[d1]-r[d2]) / r[d2] * 100) if r[d2]!=0 else None, axis=1)

    new_order = ["Field Name", m1, m2, "Missing % Change", d1, d2, "Month-to-Month % Change"]
    summary = summary[new_order]
    summary["Comment"] = ""
    summary["Approval Comments"] = ""
    return summary

def generate_distribution_df(df, analysis_type, date1):
    """
    Generate the pivot for either "value_dist" or "pop_comp" for the last 12 months,
    sorted in descending order (Dec, Nov, etc.).
    """
    months = [(date1 - relativedelta(months=i)).replace(day=1) for i in range(12)]
    months = sorted(months, reverse=True)

    sub = df[df['analysis_type'] == analysis_type].copy()
    sub['month'] = sub['filemonth_dt'].apply(lambda d: d.replace(day=1))
    sub = sub[sub['month'].isin(months)]
    grouped = sub.groupby(['field_name', 'value_label', 'month'])['value_records'].sum().reset_index()
    if grouped.empty:
        return pd.DataFrame()

    pivot = grouped.pivot_table(index=['field_name', 'value_label'],
                                columns='month', values='value_records', fill_value=0)
    pivot = pivot.reindex(columns=months, fill_value=0)

    frames = []
    for field, sub_df in pivot.groupby(level=0):
        sub_df = sub_df.droplevel(0)
        total = sub_df.sum(axis=0)
        pct = sub_df.div(total, axis=1).mul(100).round(2).fillna(0)
        data = {}
        for m in months:
            m_str = m.strftime('%Y-%m')
            data[(m_str, "Sum")] = sub_df[m]
            data[(m_str, "Percent")] = pct[m]
        tmp = pd.DataFrame(data)
        # total row
        tot_row = {}
        for m in months:
            m_str = m.strftime('%Y-%m')
            tot_row[(m_str, "Sum")] = total[m]
            tot_row[(m_str, "Percent")] = ""
        tmp.loc["Current period total"] = tot_row

        tmp.index = pd.MultiIndex.from_product([[field], tmp.index], names=["Field Name", "Value Label"])
        frames.append(tmp)

    if not frames:
        return pd.DataFrame()
    final = pd.concat(frames)
    final.columns = pd.MultiIndex.from_tuples(final.columns)
    return final

#############################################
# 3) Loading & Saving
#############################################

def load_report_data(file_path, date1, date2):
    """Load or generate Summary, ValueDist, PopComp from Excel. 
       If no 'Comment' col in those sheets, create it as empty."""
    df_data = pd.read_excel(file_path, sheet_name="Data")
    df_data["filemonth_dt"] = pd.to_datetime(df_data["filemonth_dt"])

    wb = load_workbook(file_path, data_only=True)
    # Summary
    if "Summary" in wb.sheetnames:
        summary_df = pd.read_excel(file_path, sheet_name="Summary")
    else:
        summary_df = generate_summary_df(df_data, date1, date2)

    # Value Dist
    if "Value Distribution" in wb.sheetnames:
        val_dist_df = pd.read_excel(file_path, sheet_name="Value Distribution")
    else:
        val_dist_df = generate_distribution_df(df_data, "value_dist", date1)

    # If no "Comment" column, create it
    if "Comment" not in val_dist_df.columns:
        val_dist_df["Comment"] = ""

    # Population Comparison
    if "Population Comparison" in wb.sheetnames:
        pop_comp_df = pd.read_excel(file_path, sheet_name="Population Comparison")
    else:
        pop_comp_df = generate_distribution_df(df_data, "pop_comp", date1)

    if "Comment" not in pop_comp_df.columns:
        pop_comp_df["Comment"] = ""

    return df_data, summary_df, val_dist_df, pop_comp_df

def cache_previous_comments(current_folder):
    """Scan 'previous/<folder>' for older Excel files, read summary comments from 'month to month' column, store in CSV."""
    data = []
    prev_folder = os.path.join(os.getcwd(), "previous", current_folder)
    if not os.path.exists(prev_folder):
        st.warning("Previous months folder not found.")
        return pd.DataFrame(columns=["Field Name", "Month", "Comment"])
    month_pattern = re.compile(r'(\d{4})[- ]?(\d{2})')

    for file in os.listdir(prev_folder):
        if file.lower().endswith('.xlsx'):
            file_path = os.path.join(prev_folder, file)
            m = month_pattern.search(file)
            if m:
                month_year = f"{m.group(1)}-{m.group(2)}"
            else:
                month_year = "unknown"
            try:
                wb = load_workbook(file_path, data_only=True)
            except Exception as e:
                st.error(f"Error opening previous file {file}: {e}")
                continue
            if "Summary" not in wb.sheetnames:
                continue
            ws = wb["Summary"]
            header_row = None
            for r in range(1, 4):
                headers = [cell.value for cell in ws[r] if cell.value]
                if any("month to month" in str(val).lower() for val in headers):
                    header_row = r
                    break
            if header_row is None:
                continue
            header = [cell.value for cell in ws[header_row]]
            col_index = None
            for i, col_name in enumerate(header, start=1):
                if col_name and "month to month" in str(col_name).lower():
                    col_index = i
                    break
            if col_index is None:
                continue

            for row in ws.iter_rows(min_row=header_row+1):
                field_cell = row[0]
                if field_cell.value:
                    field_name = str(field_cell.value).strip()
                    cell = row[col_index - 1]
                    comment_text = ""
                    if cell.comment and cell.comment.text:
                        raw = str(cell.comment.text).strip()
                        if raw.lower() != "nan" and raw != "":
                            comment_text = raw
                    if comment_text:
                        data.append({"Field Name": field_name, "Month": month_year, "Comment": comment_text})

    df = pd.DataFrame(data)
    df.to_csv("previous_comments.csv", index=False)
    return df

def get_cached_previous_comments(current_folder):
    """Load the CSV of previous comments if it exists, else empty."""
    try:
        df = pd.read_csv("previous_comments.csv")
        if df.empty:
            return pd.DataFrame(columns=["Field Name", "Month", "Comment"])
        return df
    except pd.errors.EmptyDataError:
        return pd.DataFrame(columns=["Field Name", "Month", "Comment"])

#############################################
# 4) Summaries & Merging
#############################################

def preserve_summary_comments(input_file_path, summary_df):
    """If 'Summary' sheet has existing 'Comment' or 'Approval Comments', keep them."""
    try:
        existing = pd.read_excel(input_file_path, sheet_name="Summary")
        comment_dict = {}
        approval_dict = {}
        if "Comment" in existing.columns:
            comment_dict = existing.set_index("Field Name")["Comment"].to_dict()
        if "Approval Comments" in existing.columns:
            approval_dict = existing.set_index("Field Name")["Approval Comments"].to_dict()
        summary_df["Comment"] = summary_df["Field Name"].map(comment_dict).fillna("")
        summary_df["Approval Comments"] = summary_df["Field Name"].map(approval_dict).fillna("")
    except Exception as e:
        st.warning(f"Could not preserve existing summary comments: {e}")
    return summary_df

def pivot_previous_comments(df, target_month):
    """Pivot only the single 'target_month' from the cached CSV."""
    if df.empty:
        return pd.DataFrame()
    df_target = df[df["Month"] == target_month]
    if df_target.empty:
        return pd.DataFrame()
    grouped = df_target.groupby("Field Name")["Comment"].apply(
        lambda x: "\n".join(x.dropna().astype(str).str.strip())
    ).reset_index()
    grouped = grouped.rename(columns={"Comment": f"comment_{target_month}"})
    return grouped

def append_prev_comment(summary_df, target_month):
    """Append the target month's comment to the 'Comment' column in Summary."""
    prev_df = get_cached_previous_comments(st.session_state.folder)
    pivot_prev = pivot_previous_comments(prev_df, target_month)
    if pivot_prev.empty:
        return summary_df

    def combine_comments(row):
        orig = row["Comment"] if pd.notna(row["Comment"]) else ""
        field = row["Field Name"]
        match = pivot_prev[pivot_prev["Field Name"] == field]
        if not match.empty:
            prev_comment = str(match.iloc[0, 1]).strip()
            if prev_comment:
                if orig:
                    return orig + "\n" + prev_comment
                else:
                    return prev_comment
        return orig

    summary_df["Comment"] = summary_df.apply(combine_comments, axis=1)
    return summary_df

#############################################
# 5) Main Streamlit App
#############################################

def main():
    st.sidebar.title("File & Date Selection")
    folder = st.sidebar.selectbox("Select Folder", ["BDCOM", "WFHMSA", "BCards"])
    folder_path = os.path.join(os.getcwd(), folder)
    st.sidebar.write(f"Folder path: {folder_path}")
    if not os.path.exists(folder_path):
        st.sidebar.error(f"Folder '{folder}' not found.")
        return

    all_files = [f for f in os.listdir(folder_path) if f.lower().endswith(('.xlsx','.xlsb'))]
    if not all_files:
        st.sidebar.error(f"No Excel files found in folder '{folder}'.")
        return

    selected_file = st.sidebar.selectbox("Select an Excel File", all_files)
    input_file_path = os.path.join(folder_path, selected_file)

    selected_date = st.sidebar.date_input("Select Date for Date1", datetime.date(2025, 1, 1))
    date1 = datetime.datetime.combine(selected_date, datetime.datetime.min.time())
    date2 = date1 - relativedelta(months=1)

    # Cache previous comments
    prev_comments_df = get_cached_previous_comments(folder)
    if prev_comments_df.empty:
        prev_comments_df = cache_previous_comments(folder)

    st.write("Cached previous comments (if multiple months, sorted by 'Month' ascending in CSV, but we can pivot in descending order):")
    st.dataframe(prev_comments_df)

    target_prev_month = (date1 - relativedelta(months=1)).strftime("%Y-%m")

    if st.sidebar.button("Generate Report"):
        df_data, summary_df, val_dist_df, pop_comp_df = load_report_data(input_file_path, date1, date2)
        # Keep old summary comments
        summary_df = preserve_summary_comments(input_file_path, summary_df)

        # Put into session state
        st.session_state.df_data = df_data
        st.session_state.summary_df = summary_df
        st.session_state.value_dist_df = val_dist_df  # has 'Comment' col
        st.session_state.pop_comp_df = pop_comp_df    # has 'Comment' col
        st.session_state.selected_file = selected_file
        st.session_state.folder = folder
        st.session_state.date1 = date1
        st.session_state.date2 = date2
        st.session_state.input_file_path = input_file_path
        st.session_state.active_field = None

    st.write("Working Directory:", os.getcwd())

    if "df_data" in st.session_state:
        st.title("FRY14M Field Analysis Summary Report")
        st.write(f"**Folder:** {st.session_state.folder}")
        st.write(f"**File:** {st.session_state.selected_file}")
        st.write(f"**Date1:** {st.session_state.date1.strftime('%Y-%m-%d')} | **Date2:** {st.session_state.date2.strftime('%Y-%m-%d')}")

        #############################################
        # 1) Value Distribution (Comment column)
        #############################################
        st.subheader("Value Distribution")
        # Flatten & normalize if needed
        st.session_state.value_dist_df = flatten_dataframe(st.session_state.value_dist_df)
        st.session_state.value_dist_df = normalize_columns(st.session_state.value_dist_df)

        val_fields = st.session_state.value_dist_df["Field Name"].unique().tolist()
        if not val_fields:
            st.warning("No Value Distribution data available.")
        else:
            active_val = st.session_state.active_field if st.session_state.active_field in val_fields else val_fields[0]
            selected_val_field = st.selectbox("Select Field (Value Dist)", val_fields,
                                              index=val_fields.index(active_val),
                                              key="val_field_select")
            st.session_state.active_field = selected_val_field

            # Filter
            filtered_val = st.session_state.value_dist_df[st.session_state.value_dist_df["Field Name"] == selected_val_field].copy()
            # Drop old "m- Comments" columns
            filtered_val = drop_prev_comments_columns(filtered_val)

            # Pivot the single target month (or you could pivot all months in descending order if you want).
            pivot_prev = pivot_previous_comments(prev_comments_df, target_prev_month)
            if not pivot_prev.empty:
                # Merge
                filtered_val = filtered_val.merge(pivot_prev,
                                                  on="Field Name",
                                                  how="left",
                                                  suffixes=('', None))
                # rename newly merged column
                old_col = f"comment_{target_prev_month}"
                if old_col in filtered_val.columns:
                    filtered_val.rename(columns={old_col: f"{target_prev_month} m- Comments"}, inplace=True)
            # Ensure "Comment" col
            if "Comment" not in filtered_val.columns:
                filtered_val["Comment"] = ""

            # Replace NaN with empty
            filtered_val = filtered_val.fillna("")

            gb_val = GridOptionsBuilder.from_dataframe(filtered_val)
            gb_val.configure_default_column(editable=True,
                                            cellStyle={'white-space':'normal','line-height':'1.2em','width':150})
            for c in filtered_val.columns:
                if c == "Comment":
                    gb_val.configure_column(c, editable=True, width=220)
                elif c.endswith("m- Comments"):
                    gb_val.configure_column(c, editable=False, width=180)
                elif c.endswith("Sum") or c.endswith("Percent"):
                    gb_val.configure_column(c, editable=False, width=120)

            val_opts = gb_val.build()
            if isinstance(val_opts, list):
                val_opts = {"columnDefs": val_opts}
            val_opts["rowSelection"] = "single"
            val_opts["pagination"] = False
            val_opts["rowHeight"] = 40
            val_opts["headerHeight"] = 80

            val_height = compute_grid_height(filtered_val, 40, 80)
            val_res = AgGrid(filtered_val,
                             gridOptions=val_opts,
                             update_mode=GridUpdateMode.VALUE_CHANGED,
                             data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
                             key="val_grid",
                             height=val_height,
                             use_container_width=True)

            updated_val = pd.DataFrame(val_res["data"]).fillna("")
            # Merge back into the big DF
            old_vdf = st.session_state.value_dist_df
            # remove rows for this field
            mask_sel = (old_vdf["Field Name"] == selected_val_field)
            st.session_state.value_dist_df = pd.concat([
                old_vdf[~mask_sel],
                updated_val
            ], ignore_index=True)

        #############################################
        # 2) Population Comparison (Comment column)
        #############################################
        st.subheader("Population Comparison")
        st.session_state.pop_comp_df = flatten_dataframe(st.session_state.pop_comp_df)
        st.session_state.pop_comp_df = normalize_columns(st.session_state.pop_comp_df)

        pop_fields = st.session_state.pop_comp_df["Field Name"].unique().tolist()
        if not pop_fields:
            st.warning("No Population Comparison data available.")
        else:
            active_pop = st.session_state.active_field if st.session_state.active_field in pop_fields else pop_fields[0]
            selected_pop_field = st.selectbox("Select Field (Pop Comp)", pop_fields,
                                              index=pop_fields.index(active_pop) if active_pop in pop_fields else 0,
                                              key="pop_field_select")
            st.session_state.active_field = selected_pop_field

            filtered_pop = st.session_state.pop_comp_df[st.session_state.pop_comp_df["Field Name"] == selected_pop_field].copy()
            filtered_pop = drop_prev_comments_columns(filtered_pop)

            pivot_prev = pivot_previous_comments(prev_comments_df, target_prev_month)
            if not pivot_prev.empty:
                filtered_pop = filtered_pop.merge(pivot_prev, 
                                                  on="Field Name",
                                                  how="left",
                                                  suffixes=('', None))
                old_col = f"comment_{target_prev_month}"
                if old_col in filtered_pop.columns:
                    filtered_pop.rename(columns={old_col: f"{target_prev_month} m- Comments"}, inplace=True)
            if "Comment" not in filtered_pop.columns:
                filtered_pop["Comment"] = ""

            filtered_pop = filtered_pop.fillna("")

            gb_pop = GridOptionsBuilder.from_dataframe(filtered_pop)
            gb_pop.configure_default_column(editable=True,
                                            cellStyle={'white-space':'normal','line-height':'1.2em','width':150})
            for c in filtered_pop.columns:
                if c == "Comment":
                    gb_pop.configure_column(c, editable=True, width=220)
                elif c.endswith("m- Comments"):
                    gb_pop.configure_column(c, editable=False, width=180)
                elif c.endswith("Sum") or c.endswith("Percent"):
                    gb_pop.configure_column(c, editable=False, width=120)

            pop_opts = gb_pop.build()
            if isinstance(pop_opts, list):
                pop_opts = {"columnDefs": pop_opts}
            pop_opts["rowSelection"] = "single"
            pop_opts["pagination"] = False
            pop_opts["rowHeight"] = 40
            pop_opts["headerHeight"] = 80

            pop_height = compute_grid_height(filtered_pop, 40, 80)
            pop_res = AgGrid(filtered_pop,
                             gridOptions=pop_opts,
                             update_mode=GridUpdateMode.VALUE_CHANGED,
                             data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
                             key="pop_grid",
                             height=pop_height,
                             use_container_width=True)

            updated_pop = pd.DataFrame(pop_res["data"]).fillna("")
            old_pdf = st.session_state.pop_comp_df
            mask_sel = (old_pdf["Field Name"] == selected_pop_field)
            st.session_state.pop_comp_df = pd.concat([
                old_pdf[~mask_sel],
                updated_pop
            ], ignore_index=True)

        #############################################
        # 3) Summary Aggregation
        #############################################
        # a) Append previous month comment into Summary
        summary_df = st.session_state.summary_df.copy()
        summary_df = append_prev_comment(summary_df, target_prev_month)
        st.session_state.summary_df = summary_df

        # b) Optionally preserve existing summary from Excel
        try:
            existing = pd.read_excel(st.session_state.input_file_path, sheet_name="Summary")
            cdict = {}
            adict = {}
            if "Comment" in existing.columns:
                cdict = existing.set_index("Field Name")["Comment"].to_dict()
            if "Approval Comments" in existing.columns:
                adict = existing.set_index("Field Name")["Approval Comments"].to_dict()
            st.session_state.summary_df["Comment"] = st.session_state.summary_df["Field Name"].map(cdict).fillna("")\
                .str.cat(st.session_state.summary_df["Comment"], sep="\n")
            st.session_state.summary_df["Approval Comments"] = st.session_state.summary_df["Field Name"].map(adict).fillna("")
        except Exception as e:
            st.warning(f"Could not preserve existing summary comments: {e}")

        # c) Aggregate from updated ValueDist & PopComp "Comment" columns
        def aggregate_current_comments():
            sum_df = st.session_state.summary_df.copy()
            for field in sum_df["Field Name"].unique():
                notes = []
                # Value Dist
                vdf = st.session_state.value_dist_df[st.session_state.value_dist_df["Field Name"]==field]
                if "Value Label" in vdf.columns and "Comment" in vdf.columns:
                    for _, row in vdf.iterrows():
                        cmt = str(row["Comment"]).strip()
                        vlb = str(row["Value Label"]).strip()
                        if cmt:
                            notes.append(f"{vlb} - {cmt}")

                # Pop Comp
                pdf = st.session_state.pop_comp_df[st.session_state.pop_comp_df["Field Name"]==field]
                if "Value Label" in pdf.columns and "Comment" in pdf.columns:
                    for _, row in pdf.iterrows():
                        cmt = str(row["Comment"]).strip()
                        vlb = str(row["Value Label"]).strip()
                        if cmt:
                            notes.append(f"{vlb} - {cmt}")

                aggregated_note = "\n".join(notes)
                if aggregated_note:
                    sum_df.loc[sum_df["Field Name"]==field, "Comment"] = aggregated_note

            st.session_state.summary_df = sum_df

        aggregate_current_comments()

        # d) Reorder summary columns so "Approval Comments" is after "Comment"
        sum_df = st.session_state.summary_df.copy()
        if "Approval Comments" not in sum_df.columns:
            sum_df["Approval Comments"] = ""
        if "Comment" not in sum_df.columns:
            sum_df["Comment"] = ""

        col_list = list(sum_df.columns)
        col_list.remove("Approval Comments")
        col_list.remove("Comment")
        new_order = ["Field Name"] + [c for c in col_list if c != "Field Name"] + ["Comment","Approval Comments"]
        sum_df = sum_df[new_order]

        st.subheader("Summary")
        comment_renderer = JsCode("function(params){return params.value ? params.value : '';}")
        gb_sum = GridOptionsBuilder.from_dataframe(sum_df)
        gb_sum.configure_default_column(editable=False, 
                                        cellStyle={'white-space':'normal','line-height':'1.2em','width':150})
        gb_sum.configure_column("Approval Comments", editable=True, width=250, minWidth=100, maxWidth=300)
        gb_sum.configure_column("Comment", editable=False, cellRenderer=comment_renderer, width=250, minWidth=100, maxWidth=300)

        for c in sum_df.columns:
            if c not in ["Field Name", "Comment", "Approval Comments"]:
                if "Change" in c:
                    gb_sum.configure_column(c, type=["numericColumn"],
                                            valueFormatter="(params.value != null ? params.value.toFixed(2)+'%' : '')",
                                            width=150, minWidth=100, maxWidth=200)
                else:
                    gb_sum.configure_column(c, type=["numericColumn"],
                                            valueFormatter="(params.value != null ? params.value.toLocaleString('en-US') : '')",
                                            width=150, minWidth=100, maxWidth=200)

        sum_opts = gb_sum.build()
        if isinstance(sum_opts, list):
            sum_opts = {"columnDefs": sum_opts}
        for c in sum_opts["columnDefs"]:
            c["headerName"] = "\n".join(c["headerName"].split())
        sum_opts["rowSelection"] = "single"
        sum_opts["pagination"] = False
        sum_opts["rowHeight"] = 40
        sum_opts["headerHeight"] = 80

        sum_height = compute_grid_height(sum_df, 40, 80)
        sum_res = AgGrid(sum_df,
                         gridOptions=sum_opts,
                         update_mode=GridUpdateMode.VALUE_CHANGED,
                         data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
                         key="sum_grid",
                         height=sum_height,
                         use_container_width=True)
        st.session_state.summary_df = pd.DataFrame(sum_res["data"])

        # Save all sheets back to Excel
        try:
            with pd.ExcelWriter(st.session_state.input_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                st.session_state.summary_df.to_excel(writer, index=False, sheet_name="Summary")
                st.session_state.value_dist_df.to_excel(writer, index=False, sheet_name="Value Distribution")
                st.session_state.pop_comp_df.to_excel(writer, index=False, sheet_name="Population Comparison")

                # Optionally store the summary comments in Excel cell comments:
                summary_sheet = writer.sheets["Summary"]
                header = [cell.value for cell in summary_sheet[1]]
                try:
                    comment_col_index = header.index("Comment") + 1
                except ValueError:
                    comment_col_index = len(header)

                for idx, row in st.session_state.summary_df.iterrows():
                    cmt = str(row["Comment"]).strip()
                    if cmt and cmt.lower()!="nan":
                        com_obj = Comment(cmt, "User")
                        summary_sheet.cell(row=idx+2, column=comment_col_index).comment = com_obj

            st.success(f"Excel file updated successfully at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}!")
        except Exception as e:
            st.error(f"Error updating the Excel file: {e}")


if __name__=="__main__":
    main()
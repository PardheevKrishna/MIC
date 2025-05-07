import pandas as pd
import datetime
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment

from dash import Dash, dcc, html, dash_table
from dash.dependencies import Input, Output

# ---------------------------
# Constants & Input/Output
# ---------------------------
INPUT_FILE = "input.xlsx"
OUTPUT_FILE = "output.xlsx"

# ---------------------------
# Step 1: Read the Excel File
# ---------------------------
df_data = pd.read_excel(INPUT_FILE, sheet_name="Data")
df_control = pd.read_excel(INPUT_FILE, sheet_name="Control")  # if you use it elsewhere

# Parse the date column
df_data['filemonth_dt'] = pd.to_datetime(df_data['filemonth_dt'], format='%m/%d/%Y')

# ---------------------------
# Step 2: Define the Two Dates
# ---------------------------
date1 = datetime.datetime(2025, 1, 1)
date2 = datetime.datetime(2024, 12, 1)

# ---------------------------
# Step 3: Unique Fields & Regex Phrases
# ---------------------------
fields = sorted(df_data['field_name'].unique())
phrases = [
    r"1\)\s*F6CF Loan - Both Pop, Diff Values",
    r"2\)\s*CF Loan - Prior Null, Current Pop",
    r"3\)\s*CF Loan - Prior Pop, Current Null"
]
def contains_phrase(text):
    text = str(text)
    return any(re.search(p, text) for p in phrases)

# ---------------------------
# Step 4: Compute Summary Data
# ---------------------------
summary_data = []
for field in fields:
    # Missing values sums
    m1 = df_data[
        (df_data['analysis_type']=='value_dist') &
        (df_data['field_name']==field) &
        (df_data['filemonth_dt']==date1)
    ]
    missing_sum_date1 = m1[m1['value_label'].str.contains("Missing", case=False, na=False)]['value_records'].sum()

    m2 = df_data[
        (df_data['analysis_type']=='value_dist') &
        (df_data['field_name']==field) &
        (df_data['filemonth_dt']==date2)
    ]
    missing_sum_date2 = m2[m2['value_label'].str.contains("Missing", case=False, na=False)]['value_records'].sum()

    # Month-to-Month diffs
    p1 = df_data[
        (df_data['analysis_type']=='pop_comp') &
        (df_data['field_name']==field) &
        (df_data['filemonth_dt']==date1)
    ]
    m2m_sum_date1 = p1[p1['value_label'].apply(contains_phrase)]['value_records'].sum()

    p2 = df_data[
        (df_data['analysis_type']=='pop_comp') &
        (df_data['field_name']==field) &
        (df_data['filemonth_dt']==date2)
    ]
    m2m_sum_date2 = p2[p2['value_label'].apply(contains_phrase)]['value_records'].sum()

    summary_data.append([
        field,
        missing_sum_date1,
        missing_sum_date2,
        m2m_sum_date1,
        m2m_sum_date2,
        ""  # Initialize the comment column with empty strings (will be used to store comments)
    ])

# Build a Pandas DataFrame for Dash
df_summary = pd.DataFrame(summary_data, columns=[
    "Field Name",
    f"Missing {date1.strftime('%m/%d/%Y')}",
    f"Missing {date2.strftime('%m/%d/%Y')}",
    f"M2M Diff {date1.strftime('%m/%d/%Y')}",
    f"M2M Diff {date2.strftime('%m/%d/%Y')}",
    "Comment",  # New column for comment storage
])

# ---------------------------
# Step 5: Build Value-Dist & Pop-Comp DataFrames
# ---------------------------
# filter to just the two months
mask_months = df_data['filemonth_dt'].isin([date1, date2])

df_value_dist = (
    df_data[mask_months & (df_data['analysis_type']=='value_dist')]
    .copy()
)
df_value_dist['filemonth_dt'] = df_value_dist['filemonth_dt'].dt.strftime('%m/%d/%Y')

df_pop_comp = (
    df_data[mask_months & (df_data['analysis_type']=='pop_comp')]
    .loc[lambda d: d['value_label'].apply(contains_phrase)]
    .copy()
)
df_pop_comp['filemonth_dt'] = df_pop_comp['filemonth_dt'].dt.strftime('%m/%d/%Y')

# ---------------------------
# Step 6: Generate & Style the Excel Report
# ---------------------------
wb = load_workbook(INPUT_FILE)

# Remove existing Summary sheet if it exists
if "Summary" in wb.sheetnames:
    wb.remove(wb["Summary"])

ws = wb.create_sheet("Summary")

# Styles
header_fill = PatternFill("solid", fgColor="4F81BD")
header_font = Font(bold=True, color="FFFFFF")
center = Alignment(horizontal="center", vertical="center")
white_fill = PatternFill("solid", fgColor="FFFFFF")
gray_fill  = PatternFill("solid", fgColor="D3D3D3")
thick = Side(border_style="thick", color="000000")
thick_border = Border(thick, thick, thick, thick)

# Title row
ws.merge_cells("A1:F1")
cell = ws["A1"]
cell.value = "BDCOMM FRY14M Field Analysis Summary"
cell.fill = header_fill
cell.font = header_font
cell.alignment = center

# Column headers rows 2–3
ws.merge_cells("A2:A3")
h = ws["A2"]
h.value, h.fill, h.font, h.alignment = "Field Name", header_fill, header_font, center

ws.merge_cells("C2:D2")
h = ws["C2"]
h.value, h.fill, h.font, h.alignment = "Missing Values", header_fill, header_font, center

ws["C3"].value = date1
ws["C3"].number_format = "mm/dd/yyyy"
ws["C3"].alignment = center
ws["D3"].value = date2
ws["D3"].number_format = "mm/dd/yyyy"
ws["D3"].alignment = center

ws.merge_cells("F2:G2")
h = ws["F2"]
h.value, h.fill, h.font, h.alignment = "Month to Month Value Differences", header_fill, header_font, center

ws["F3"].value = date1
ws["F3"].number_format = "mm/dd/yyyy"
ws["F3"].alignment = center
ws["G3"].value = date2
ws["G3"].number_format = "mm/dd/yyyy"
ws["G3"].alignment = center

ws.merge_cells("I2:I3")
h = ws["I2"]
h.value, h.fill, h.font, h.alignment = "Approval Comments", header_fill, header_font, center

# Write data starting at row 4
start_row = 4
for i, row in enumerate(summary_data, start=start_row):
    fill = white_fill if (i - start_row) % 2 == 0 else gray_fill

    # Field Name
    c = ws.cell(row=i, column=1, value=row[0])
    c.fill, c.alignment, c.border = fill, center, thick_border

    # Missing date1 & date2
    for j, val in enumerate(row[1:3], start=3):
        c = ws.cell(row=i, column=j, value=val)
        c.fill, c.alignment, c.border = fill, center, thick_border

    # M2M date1 & date2
    for j, val in enumerate(row[3:], start=6):
        c = ws.cell(row=i, column=j, value=val)
        c.fill, c.alignment, c.border = fill, center, thick_border

    # Comment (New column) - Only if there's a comment
    c = ws.cell(row=i, column=6, value=row[5])
    c.fill, c.alignment, c.border = fill, center, thick_border

# Auto-width
for col in list("ABCDEF"):
    ws.column_dimensions[col].width = 20

wb.save(OUTPUT_FILE)

# ---------------------------
# Step 7: Dash App with Three Tabs
# ---------------------------
app = Dash(__name__)

app.layout = html.Div([
    html.H2("BDCOMM FRY14M Field Analysis"),
    dcc.Tabs([
        dcc.Tab(label="Summary", children=[
            dash_table.DataTable(
                id='summary_table',
                columns=[{"name": c, "id": c} for c in df_summary.columns if c != 'Comment'],  # Remove comment column
                data=df_summary.to_dict("records"),
                page_size=20,
                style_header={'backgroundColor': '#4F81BD', 'color': 'white', 'fontWeight': 'bold'},
                style_cell={'textAlign': 'center'},
                style_table={'overflowX': 'auto'},
                selected_cells=[],  # For capturing double-clicks
                editable=False,  # Summary table is not editable
                style_data_conditional=[
                    {
                        'if': {'column_id': c, 'filter_query': '{Comment} != ""'},
                        'backgroundColor': 'yellow',
                    }
                    for c in ['C', 'D', 'F', 'G']  # Highlight columns with comments (Missing, M2M, etc.)
                ]
            )
        ]),
        dcc.Tab(label="Value Distribution", children=[
            html.Div(id='value_sql_logic_box', style={'padding': '10px', 'backgroundColor': '#f5f5f5'}),
            dash_table.DataTable(
                id='value_dist_table',
                columns=[{"name": c, "id": c} for c in df_value_dist.columns if c != 'value_sql_logic'],  # Remove the column
                data=df_value_dist.to_dict("records"),
                page_size=20,
                style_header={'fontWeight': 'bold', 'backgroundColor': '#4F81BD', 'color': 'white'},
                style_cell={'textAlign': 'left'},
                style_table={'overflowX': 'auto'},
                editable=True,  # Enable comment editing
            )
        ]),
        dcc.Tab(label="Population Comparison", children=[
            html.Div(id='pop_sql_logic_box', style={'padding': '10px', 'backgroundColor': '#f5f5f5'}),
            dash_table.DataTable(
                id='pop_comp_table',
                columns=[{"name": c, "id": c} for c in df_pop_comp.columns if c != 'value_sql_logic'],  # Remove the column
                data=df_pop_comp.to_dict("records"),
                page_size=20,
                style_header={'fontWeight': 'bold', 'backgroundColor': '#4F81BD', 'color': 'white'},
                style_cell={'textAlign': 'left'},
                style_table={'overflowX': 'auto'},
                editable=True,  # Enable comment editing
            )
        ]),
    ])
])


# Callback for handling comment updates in Value Distribution and Population Comparison
@app.callback(
    [Output('summary_table', 'data'),
     Output('value_sql_logic_box', 'children'),
     Output('pop_sql_logic_box', 'children')],
    [Input('value_dist_table', 'data'),
     Input('pop_comp_table', 'data')]
)
def update_comments(value_dist_data, pop_comp_data):
    # Update the Summary Table with new comments
    if value_dist_data:
        updated_value_dist = pd.DataFrame(value_dist_data)
        for i, row in updated_value_dist.iterrows():
            df_summary.loc[df_summary['Field Name'] == row['field_name'], 'Comment'] = row.get('comment', '')

    if pop_comp_data:
        updated_pop_comp = pd.DataFrame(pop_comp_data)
        for i, row in updated_pop_comp.iterrows():
            df_summary.loc[df_summary['Field Name'] == row['field_name'], 'Comment'] = row.get('comment', '')

    # Save changes back to Excel
    wb = load_workbook(OUTPUT_FILE)
    ws = wb['Summary']
    for idx, row in df_summary.iterrows():
        comment = row['Comment']
        if comment:  # If a comment exists, add it as a cell comment
            cell = ws.cell(row=idx+4, column=6)  # Assuming the Comment is in the 6th column
            cell.comment = Comment(comment, "User")

    wb.save(OUTPUT_FILE)

    return df_summary.to_dict("records"), "", ""


# Run the app
if __name__ == "__main__":
    app.run(debug=True)
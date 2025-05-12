#!/usr/bin/env python3
import os
import glob
import csv
from datetime import datetime
from openpyxl import load_workbook

def extract_comments_from_folder(folder_path, output_csv='previous_comments.csv'):
    # prepare output file
    with open(output_csv, 'w', newline='', encoding='utf-8') as f_out:
        writer = csv.writer(f_out)
        writer.writerow(['field_name', 'date', 'research', 'comment'])

        # find all .xlsx files under folder_path
        pattern = os.path.join(folder_path, '*.xlsx')
        for filepath in glob.glob(pattern):
            wb = load_workbook(filepath, data_only=True)
            if 'Summary' not in wb.sheetnames:
                continue
            ws = wb['Summary']

            # scan every cell for a comment
            for row in ws.iter_rows():
                for cell in row:
                    if cell.comment is None:
                        continue

                    row_idx    = cell.row
                    col_letter = cell.column_letter

                    # 1) field_name is in column A of the same row
                    field_name = ws[f'A{row_idx}'].value

                    # 2) date is in the 3rd row of the same column
                    raw_date = ws[f'{col_letter}3'].value
                    # normalize to mm/dd/yyyy
                    if isinstance(raw_date, datetime):
                        date_str = raw_date.strftime('%m/%d/%Y')
                    else:
                        # attempt parse if it’s a string
                        date_obj = datetime.strptime(str(raw_date), '%m/%d/%Y')
                        date_str = date_obj.strftime('%m/%d/%Y')

                    # 3) research tag depends on column
                    if col_letter in ('C', 'D'):
                        research = 'Missing'
                    elif col_letter in ('F', 'G'):
                        research = 'M2M Diff'
                    else:
                        research = ''

                    # 4) the comment text itself
                    comment_text = cell.comment.text.strip()

                    writer.writerow([field_name, date_str, research, comment_text])

    print(f'Done! Extracted comments saved to {output_csv}')

if __name__ == '__main__':
    folder = '/path/to/your/excel/folder'   # ← change me
    extract_comments_from_folder(folder)
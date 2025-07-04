#!/usr/bin/env python3
import os
import getpass
import time
import threading
import logging
import datetime
import tkinter as tk
from tkinter import filedialog, ttk, messagebox

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import xlsxwriter
from tqdm import tqdm
from PyPDF2 import PdfReader, PdfWriter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfgen.canvas import Canvas

# ─── Logging ───────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s: %(message)s",
    datefmt="%H:%M:%S"
)

# ─── Tk UI ─────────────────────────────────────────────────────────────────
root = tk.Tk()
root.title("Metrics ⇒ PDF + Excel")
frame = ttk.Frame(root, padding=10)
frame.grid()

btn = ttk.Button(frame, text="Upload Excel")
btn.grid(row=0, column=0)
lbl_file = ttk.Label(frame, text="No file selected")
lbl_file.grid(row=0, column=1, padx=10)

lbl_by   = ttk.Label(frame, text="Uploaded by: N/A")
lbl_date = ttk.Label(frame, text="Uploaded date: N/A")
lbl_time = ttk.Label(frame, text="Process time: N/A")
lbl_rows = ttk.Label(frame, text="Processed rows: 0")
for i, w in enumerate((lbl_by, lbl_date, lbl_time, lbl_rows), start=1):
    w.grid(row=i, column=0, columnspan=2, sticky="w", pady=2)

def update_progress(r, elapsed):
    lbl_rows.config(text=f"Processed rows: {r}")
    lbl_time.config(text=f"Elapsed time: {elapsed:.2f}s")

# ─── PDF Canvas for per-page tqdm ────────────────────────────────────────────
class ProgressCanvas(Canvas):
    def __init__(self, filename, progress, **kwargs):
        super().__init__(filename, **kwargs)
        self._pbar = progress
    def showPage(self):
        super().showPage()
        self._pbar.update(1)
    def save(self):
        super().save()
        self._pbar.close()

# ─── Main worker ────────────────────────────────────────────────────────────
def process_file(path):
    t0 = time.perf_counter()
    user    = getpass.getuser()
    today   = datetime.date.today().strftime("%m/%d/%Y")
    datefn  = datetime.date.today().strftime("%m-%d-%Y")
    input_nm = os.path.splitext(os.path.basename(path))[0]

    # UI: uploaded by/date
    root.after(0, lambda: lbl_by.config(text=f"Uploaded by: {user}"))
    root.after(0, lambda: lbl_date.config(text=f"Uploaded date: {today}"))

    # 1) Header row
    logging.info("Reading header row…")
    wb_h = load_workbook(path, read_only=True)
    ws_h = wb_h.active
    hdr = [c.value for c in ws_h[1]][:22]
    wb_h.close()
    colA_name = hdr[0]
    val_names  = hdr[1:11]
    var_names  = hdr[12:22]

    # 2) Stream‐read rows
    logging.info("Streaming rows from Excel…")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    total_rows = ws.max_row - 1

    colA             = np.empty(total_rows, dtype="U50")
    val_block        = np.full((total_rows,10), np.nan)
    var_input_block  = np.full((total_rows,10), np.nan)
    var_block        = np.full((total_rows,10), np.nan)  # final

    it = ws.iter_rows(
        min_row=2, max_row=ws.max_row,
        min_col=1, max_col=22, values_only=True
    )
    for i, row in enumerate(tqdm(it, total=total_rows,
                                  desc="Reading Excel", unit="row")):
        a = row[0]
        colA[i] = (a.strftime("%m/%d/%Y")
                   if isinstance(a, (datetime.date, datetime.datetime))
                   else (str(a) if a is not None else ""))
        val_block[i,:]       = row[1:11]
        var_input_block[i,:] = row[12:22]
        elapsed = time.perf_counter() - t0
        root.after(0, update_progress, i+1, elapsed)
    wb.close()
    logging.info(f"Read {total_rows:,} rows in {time.perf_counter()-t0:.2f}s")

    # 3) Build var_block: preserve input, fill gaps skipping nulls
    logging.info("Computing variances skipping nulls…")
    var_block[:] = var_input_block  # start with user-supplied
    for j in tqdm(range(10), desc="Variance calc", unit="metric"):
        v = val_block[:, j]
        idx = np.where(~np.isnan(v))[0]
        for k in range(len(idx)-1):
            i0, i1 = idx[k], idx[k+1]
            if np.isnan(var_input_block[i0, j]):
                var_block[i0, j] = v[i0] - v[i1]

    # 4) Summary stats + Count/Mean/AbsMean
    logging.info("Computing summary statistics…")
    stats = [
        "Q0","Q1","Q2","Q3","Q4","IQR",
        "Lower 2SD","Upper 2SD",
        "Lower 3SD","Upper 3SD",
        "Lower 4SD","Upper 4SD",
        "Rec Lower Thresh","Rec Upper Thresh",
        "Count","Mean","Absolute Mean"
    ]
    df_stats = pd.DataFrame(index=stats, columns=var_names, dtype=float)

    for j,mn in tqdm(enumerate(var_names),
                    total=len(var_names),
                    desc="Summary stats", unit="metric"):
        # choose input series for Count
        inp = var_input_block[:,j]
        if np.isnan(inp).all():
            # no input var's: count is number of non-nulls in val
            inp = val_block[:,j]

        nonnan = inp[~np.isnan(inp)]
        cnt    = nonnan.size
        meanv  = nonnan.mean() if cnt else 0.0
        absmean= np.mean(np.abs(nonnan)) if cnt else 0.0

        # fill stats if we have computed var_block
        vb = var_block[:, j]
        vn = vb[~np.isnan(vb)]
        if vn.size:
            q = np.percentile(vn, [0,25,50,75,100])
            sd = vn.std(ddof=1)
            df_stats.loc[["Q0","Q1","Q2","Q3","Q4"], mn] = q
            df_stats.at["IQR", mn] = q[3] - q[1]
            for k in (2,3,4):
                df_stats.at[f"Lower {k}SD", mn] = meanv - k*sd
                df_stats.at[f"Upper {k}SD", mn] = meanv + k*sd
            df_stats.at["Rec Lower Thresh", mn] = round(meanv - 3*sd, -3)
            df_stats.at["Rec Upper Thresh", mn] = round(meanv + 3*sd, -3)

        df_stats.at["Count", mn]         = cnt
        df_stats.at["Mean", mn]          = meanv
        df_stats.at["Absolute Mean", mn] = absmean

    # 5) Write Excel row-by-row
    out_xlsx = f"ThresholdAnalysis_output_{input_nm}_{datefn}.xlsx"
    logging.info(f"Writing '{out_xlsx}'…")
    wb_x = xlsxwriter.Workbook(out_xlsx, {
        'constant_memory': True,
        'nan_inf_to_errors': True
    })
    ws_x = wb_x.add_worksheet()
    hdr_fmt = wb_x.add_format({'bold': True})
    ws_x.set_row(0, None, hdr_fmt)
    ws_x.freeze_panes(1,1)
    ws_x.write_row(0, 0, [colA_name] + var_names)
    for i in tqdm(range(total_rows),
                  desc="Writing Excel rows", unit="row"):
        rowv = [colA[i]] + [
            var_block[i,j] if not np.isnan(var_block[i,j]) else None
            for j in range(10)
        ]
        ws_x.write_row(i+1, 0, rowv)
    wb_x.close()

    # Auto-fit via openpyxl
    logging.info("Auto-fitting columns…")
    wb2 = load_workbook(out_xlsx)
    ws2 = wb2.active
    sample_n = min(1000, total_rows)
    for col in tqdm(ws2.columns,
                    desc="Auto-fitting cols", unit="col"):
        vals = [cell.value for cell in col[:sample_n+1]]
        maxl = max(len(str(v)) if v is not None else 0 for v in vals)
        ws2.column_dimensions[col[0].column_letter].width = maxl + 2
    wb2.save(out_xlsx)

    # 6) Build one-page PDF summary + hyperlink
    out_pdf = f"output_summary_{datefn}.pdf"
    logging.info(f"Building '{out_pdf}'…")
    doc = SimpleDocTemplate(
        out_pdf,
        pagesize=landscape(A4),
        leftMargin=20, rightMargin=20,
        topMargin=20, bottomMargin=20
    )
    styles = getSampleStyleSheet()
    elems = []

    # Meta info
    for lbl,val in [
        ("Uploaded by", user),
        ("Uploaded date", today),
        ("Original file", os.path.basename(path)),
        ("Process time", f"{time.perf_counter()-t0:.2f}s")
    ]:
        elems.append(Paragraph(f"<b>{lbl}:</b> {val}", styles["Normal"]))
    elems.append(Spacer(1,12))

    # Summary table data
    header_row = ["Statistic"] + var_names
    data = [header_row] + [
        [stat] + [f"{df_stats.at[stat,m]:,.2f}" for m in var_names]
        for stat in stats
    ]
    last2 = len(data)-1
    last1 = last2-1

    # fix column widths
    page_w,_ = landscape(A4)
    avail_w  = page_w - doc.leftMargin - doc.rightMargin
    cw       = avail_w / len(header_row)

    tbl_style = TableStyle([
        ("GRID",(0,0),(-1,-1),0.25,colors.black),
        ("BACKGROUND",(0,0),(-1,0),colors.lightgrey),
        ("ALIGN",(1,0),(-1,-1),"RIGHT"),
        ("FONTSIZE",(0,0),(-1,0),10),
        ("FONTSIZE",(0,1),(-1,-1),8),
        ("FONTNAME",(0,last1),(-1,last1),"Helvetica-Bold"),
        ("FONTNAME",(0,last2),(-1,last2),"Helvetica-Bold"),
    ])
    elems.append(Table(
        data,
        repeatRows=1,
        hAlign="LEFT",
        colWidths=[cw]*len(header_row),
        style=tbl_style
    ))
    elems.append(Spacer(1,12))

    # hyperlink
    link = f"file:///{os.path.abspath(out_xlsx)}"
    elems.append(Paragraph(
        f'<a href="{link}">Download full variances Excel</a>',
        styles["Normal"]
    ))

    pbar_pdf = tqdm(total=1, desc="PDF pages", unit="page")
    doc.build(
        elems,
        canvasmaker=lambda fn, **kw:
            ProgressCanvas(fn, progress=pbar_pdf, **kw)
    )

    # embed the Excel into PDF
    with open(out_xlsx, "rb") as xf, open(out_pdf, "rb") as pf:
        reader = PdfReader(pf)
        writer = PdfWriter()
        writer.append_pages_from_reader(reader)
        writer.add_attachment(os.path.basename(out_xlsx), xf.read())
        with open(out_pdf, "wb") as outf:
            writer.write(outf)

    # final UI update
    def finish():
        elapsed = time.perf_counter() - t0
        lbl_time .config(text=f"Process time: {elapsed:.2f}s")
        lbl_rows .config(text=f"Processed rows: {total_rows}")
        messagebox.showinfo("Done",
            f"Finished in {elapsed:.2f}s\nPDF → {out_pdf}\nExcel → {out_xlsx}"
        )
        btn.config(state="normal")
    root.after(0, finish)

# ─── Hookup ────────────────────────────────────────────────────────────────
def on_upload():
    fn = filedialog.askopenfilename(
        title="Select Excel file",
        filetypes=[("Excel","*.xlsx *.xls")]
    )
    if not fn:
        return
    lbl_file.config(text=os.path.basename(fn))
    btn.config(state="disabled")
    threading.Thread(target=process_file,
                     args=(fn,), daemon=True).start()

btn.config(command=on_upload)
root.mainloop()
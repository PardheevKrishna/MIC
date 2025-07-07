import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

def clean_excel_to_dataframe(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # Fill merged cells with their top-left value
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged_range.bounds
        value = ws.cell(row=min_row, column=min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if not isinstance(cell, Cell) or (row == min_row and col == min_col):
                    continue
                ws.cell(row=row, column=col)._value = value  # safe bypass

    # Extract as grid
    data = [[cell.value for cell in row] for row in ws.iter_rows()]

    # Convert to DataFrame
    df = pd.DataFrame(data)
    df.dropna(how='all', inplace=True)
    df.dropna(axis=1, how='all', inplace=True)
    df.reset_index(drop=True, inplace=True)

    # Use row with most non-null values as header
    header_idx = df.notnull().sum(axis=1).idxmax()
    df.columns = df.iloc[header_idx]
    df = df[header_idx + 1:].reset_index(drop=True)

    return df.convert_dtypes()

# Usage
df_clean = clean_excel_to_dataframe("your_file.xlsx")
print(df_clean.head())
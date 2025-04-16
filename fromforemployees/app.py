import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
import os
import logging
from openpyxl.styles import PatternFill

# Set up logging
logging.basicConfig(level=logging.INFO)

# Files
LOGS_FILE = 'logs.xlsx'
LEAVES_FILE = 'leaves.xlsx'

# Predefined list of employees for TM dropdown
EMPLOYEES = ["Alice", "Bob", "Carol", "Dave"]

# Orange fill for anomaly rows
ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

# ---------------------------------------------------
# Load suggestions for dropdown fields from all employee sheets.
# It will scan through each sheet in LOGS_FILE whose name is one of EMPLOYEES.
# For each row in these sheets, extract values from specific columns:
#
#   Column 2: "Main project"
#   Column 3: "Name of the Project"
#   Column 4: "Project Key Milestones"
#   Column 8: "Status"
#
# If a sheet does not exist for an employee, it is skipped.
def load_suggestions():
    suggestions = {
        "main_project": set(),
        "project_name": set(),
        "project_key_milestones": set(),
        "status": set()
    }
    if os.path.exists(LOGS_FILE):
        try:
            wb = openpyxl.load_workbook(LOGS_FILE)
            for sheet in wb.worksheets:
                # Only process sheets with names in EMPLOYEES
                if sheet.title not in EMPLOYEES:
                    continue
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Check length for safety then extract:
                    # Column 2: index 1 => Main project
                    if len(row) >= 2 and row[1]:
                        suggestions["main_project"].add(str(row[1]).strip())
                    # Column 3: index 2 => Name of the Project
                    if len(row) >= 3 and row[2]:
                        suggestions["project_name"].add(str(row[2]).strip())
                    # Column 4: index 3 => Project Key Milestones
                    if len(row) >= 4 and row[3]:
                        suggestions["project_key_milestones"].add(str(row[3]).strip())
                    # Column 8: index 7 => Status
                    if len(row) >= 8 and row[7]:
                        suggestions["status"].add(str(row[7]).strip())
        except Exception as e:
            logging.error("Error loading suggestions: %s", e)
    # Sort lists for each field:
    for key in suggestions:
        suggestions[key] = sorted(list(suggestions[key]))
    return suggestions

# ---------------------------------------------------
# This function displays both a text input and a selectbox.
# Users can type a custom value or choose one from the dropdown.
# If the text input is not empty, that value is given priority.
def handle_input_or_select(field_name, suggestions):
    custom_value = st.text_input(f"Enter {field_name} (or select from dropdown):", key=f"{field_name}_input")
    # Include an empty string at the top of the dropdown so that if nothing is selected,
    # the text input (if provided) will be used.
    selected_value = st.selectbox(f"Select {field_name}", options=[""] + suggestions, key=f"{field_name}_select")
    return custom_value.strip() if custom_value.strip() else selected_value

# ---------------------------------------------------
# Validate the provided data.
# Status Date can be any valid date.
def validate_fields(data):
    errors = []
    try:
        datetime.strptime(data['status_date'], "%Y-%m-%d")
    except ValueError:
        errors.append("Invalid Status Date format.")
        
    if not data['main_project'].strip():
        errors.append("Main project is required.")
    if not data['project_name'].strip():
        errors.append("Name of the Project is required.")
    if not data['project_key_milestones'].strip():
        errors.append("Project Key Milestones is required.")
    if not data['tm'].strip():
        errors.append("Team Member (TM) is required.")
        
    try:
        datetime.strptime(data['start_date'], "%Y-%m-%d")
    except ValueError:
        errors.append("Invalid Start Date format.")
    try:
        completion_date = datetime.strptime(data['completion_date'], "%Y-%m-%d").date()
        start_date = datetime.strptime(data['start_date'], "%Y-%m-%d").date()
        if start_date > completion_date:
            errors.append("Start Date cannot be after Completion Date.")
    except ValueError:
        errors.append("Invalid Completion Date format.")
            
    try:
        pct = float(data['percent_completion'])
        if pct < 0 or pct > 100:
            errors.append("% of Completion must be between 0 and 100.")
    except ValueError:
        errors.append("Invalid % of Completion.")
        
    if not data['status'].strip():
        errors.append("Status is required.")
        
    try:
        wts = float(data['weekly_time_spent'])
        if wts < 0:
            errors.append("Weekly Time Spent cannot be negative.")
    except ValueError:
        errors.append("Invalid Weekly Time Spent value.")
        
    try:
        ph = float(data['projected_hours'])
        if ph < 0:
            errors.append("Projected Hours cannot be negative.")
    except ValueError:
        errors.append("Invalid Projected Hours value.")
        
    if not data['functional_area'].strip():
        errors.append("Functional Area is required.")
    if not data['project_category'].strip():
        errors.append("Project Category is required.")
    if data['complexity'] not in ["H", "M", "L"]:
        errors.append("Complexity must be H, M, or L.")
    if data['novelty'] not in ["BAU repetitive", "One time repetitive", "New one time"]:
        errors.append("Novelity must be one of BAU repetitive, One time repetitive, or New one time.")
    if not data['output_type'].strip():
        errors.append("Output Type is required.")
    if not data['impact_type'].strip():
        errors.append("Impact type is required.")
    
    return errors

# ---------------------------------------------------
# Count leave days for a TM based on the week (Monday to Friday) that includes the given status_date.
def count_leave_days(tm, status_date):
    leave_count = 0
    if os.path.exists(LEAVES_FILE):
        try:
            wb = openpyxl.load_workbook(LEAVES_FILE)
            sheet = wb.active
            current_date = datetime.strptime(status_date, "%Y-%m-%d").date()
            monday = current_date - timedelta(days=current_date.weekday())
            friday = monday + timedelta(days=4)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if str(row[0]).strip() == tm:
                    try:
                        leave_date = datetime.strptime(str(row[1]).strip(), "%Y-%m-%d").date()
                        if monday <= leave_date <= friday:
                            leave_count += 1
                    except:
                        continue
        except Exception as e:
            logging.error("Error reading leaves.xlsx: %s", e)
    return leave_count

# ---------------------------------------------------
# Check for anomalies based on Weekly Time Spent.
def check_anomalies(data):
    anomalies = []
    try:
        current_date = datetime.strptime(data['status_date'], "%Y-%m-%d").date()
        weekly_time_spent = float(data['weekly_time_spent'])
    except:
        return anomalies
    tm = data['tm'].strip()
    leave_days = count_leave_days(tm, data['status_date'])
    allowed_days = 5 - leave_days
    max_allowed_hours = allowed_days * 9
    if weekly_time_spent > max_allowed_hours:
        anomalies.append(f"Weekly Time Spent exceeds allowed limit. With {leave_days} leave day(s), maximum allowed is {max_allowed_hours} hrs.")
    return anomalies

# ---------------------------------------------------
# Append the log to the employee-specific sheet.
# If the sheet for the TM does not exist, it is created with headers.
def append_log(data, anomaly_message):
    try:
        if os.path.exists(LOGS_FILE):
            wb = openpyxl.load_workbook(LOGS_FILE)
        else:
            wb = openpyxl.Workbook()
            # Remove the default sheet if it exists.
            default_sheet = wb.active
            wb.remove(default_sheet)
            
        sheet_name = data['tm']
        if sheet_name not in wb.sheetnames:
            sheet = wb.create_sheet(title=sheet_name)
            headers = [
                "Status Date (Every Friday)", "Main project", "Name of the Project", "Project Key Milestones", "TM",
                "Start Date", "Completion Date % of Completion", "Status", "Weekly Time Spent(Hrs)",
                "Projected hours (Based on the Project: End to End implementation)",
                "Functional Area (CRIT, CRIT - Data Management, CRIT - Data Governance, CRIT - Regulatory Reporting, CRIT - Portfolio Reporting, CRIT - Transformation)",
                "Project Category (Data Infrastructure, Monitoring & Insights, Analytics / Strategy Development, GDA Related, Trainings and Team Meeting)",
                "Complexity (H,M,L)", 
                "Novelity (BAU repetitive, One time repetitive, New one time)",
                "Output Type (Core production work, Ad-hoc long-term projects, Ad-hoc short-term projects, Business Management, Administration, Trainings/L&D activities, Others)",
                "Impact type (Customer Experience, Financial impact, Insights, Risk reduction, Others)",
                "Anomaly Reason"
            ]
            sheet.append(headers)
        else:
            sheet = wb[sheet_name]
            
        new_row = sheet.max_row + 1
        sheet.cell(row=new_row, column=1, value=data['status_date'])
        sheet.cell(row=new_row, column=2, value=data['main_project'])
        sheet.cell(row=new_row, column=3, value=data['project_name'])
        sheet.cell(row=new_row, column=4, value=data['project_key_milestones'])
        sheet.cell(row=new_row, column=5, value=data['tm'])
        sheet.cell(row=new_row, column=6, value=data['start_date'])
        sheet.cell(row=new_row, column=7, value=data['completion_date'])
        cell_pct = sheet.cell(row=new_row, column=8)
        cell_pct.value = float(data['percent_completion'])
        cell_pct.number_format = '0.00'
        sheet.cell(row=new_row, column=9, value=data['status'])
        cell_wts = sheet.cell(row=new_row, column=10)
        cell_wts.value = float(data['weekly_time_spent'])
        cell_wts.number_format = '0.00'
        cell_ph = sheet.cell(row=new_row, column=11)
        cell_ph.value = float(data['projected_hours'])
        cell_ph.number_format = '0.00'
        sheet.cell(row=new_row, column=12, value=data['functional_area'])
        sheet.cell(row=new_row, column=13, value=data['project_category'])
        sheet.cell(row=new_row, column=14, value=data['complexity'])
        sheet.cell(row=new_row, column=15, value=data['novelty'])
        sheet.cell(row=new_row, column=16, value=data['output_type'])
        sheet.cell(row=new_row, column=17, value=data['impact_type'])
        sheet.cell(row=new_row, column=18, value=anomaly_message)
        
        if anomaly_message:
            for col in range(1, 19):
                sheet.cell(row=new_row, column=col).fill = ORANGE_FILL
                
        wb.save(LOGS_FILE)
        logging.info("Log appended successfully for %s.", data['tm'])
        return True, None
    except Exception as e:
        logging.error("Error appending log: %s", e)
        return False, str(e)

# ---------------------------------------------------
# STREAMLIT UI
st.set_page_config(page_title="Project Log", layout="wide")
st.title("Project Log")

# Load dropdown suggestions from all employee sheets
suggestions = load_suggestions()

with st.form(key="project_log_form"):
    status_date = st.date_input("Status Date")
    
    # Handle fields with dynamic suggestions (custom input or dropdown)
    main_project = handle_input_or_select("Main project", suggestions["main_project"])
    project_name = handle_input_or_select("Name of the Project", suggestions["project_name"])
    project_key_milestones = handle_input_or_select("Project Key Milestones", suggestions["project_key_milestones"])
    
    # For Team Member, use predefined list.
    tm = st.selectbox("Team Member (TM)", options=EMPLOYEES)
    
    # Other fields
    start_date = st.date_input("Start Date")
    completion_date = st.date_input("Completion Date")
    percent_completion = st.number_input("% of Completion", min_value=0, max_value=100)
    status = st.selectbox("Status", options=[""] + suggestions["status"])
    weekly_time_spent = st.number_input("Weekly Time Spent (Hrs)", min_value=0.0, step=0.5)
    projected_hours = st.number_input("Projected hours (Based on the Project: End to End implementation)", min_value=0.0, step=0.5)
    functional_area = st.selectbox("Functional Area (CRIT, CRIT - Data Management, CRIT - Data Governance, CRIT - Regulatory Reporting, CRIT - Portfolio Reporting, CRIT - Transformation)",
                                   options=["CRIT", "CRIT - Data Management", "CRIT - Data Governance",
                                            "CRIT - Regulatory Reporting", "CRIT - Portfolio Reporting", "CRIT - Transformation"])
    project_category = st.selectbox("Project Category (Data Infrastructure, Monitoring & Insights, Analytics / Strategy Development, GDA Related, Trainings and Team Meeting)",
                                    options=["Data Infrastructure", "Monitoring & Insights", "Analytics / Strategy Development", "GDA Related", "Trainings and Team Meeting"])
    complexity = st.selectbox("Complexity (H,M,L)", options=["H", "M", "L"])
    novelty = st.selectbox("Novelity (BAU repetitive, One time repetitive, New one time)",
                           options=["BAU repetitive", "One time repetitive", "New one time"])
    output_type = st.selectbox("Output Type (Core production work, Ad-hoc long-term projects, Ad-hoc short-term projects, Business Management, Administration, Trainings/L&D activities, Others)",
                               options=["Core production work", "Ad-hoc long-term projects", "Ad-hoc short-term projects", "Business Management", "Administration", "Trainings/L&D activities", "Others"])
    impact_type = st.selectbox("Impact type (Customer Experience, Financial impact, Insights, Risk reduction, Others)",
                               options=["Customer Experience", "Financial impact", "Insights", "Risk reduction", "Others"])

    submit_button = st.form_submit_button(label="Submit")

if submit_button:
    data = {
        "status_date": str(status_date),
        "main_project": main_project,
        "project_name": project_name,
        "project_key_milestones": project_key_milestones,
        "tm": tm,
        "start_date": str(start_date),
        "completion_date": str(completion_date),
        "percent_completion": str(percent_completion),
        "status": status,
        "weekly_time_spent": str(weekly_time_spent),
        "projected_hours": str(projected_hours),
        "functional_area": functional_area,
        "project_category": project_category,
        "complexity": complexity,
        "novelty": novelty,
        "output_type": output_type,
        "impact_type": impact_type
    }

    # Validate the inputs
    field_errors = validate_fields(data)
    if field_errors:
        st.error(" ".join(field_errors))
    else:
        anomalies = check_anomalies(data)
        anomaly_message = "; ".join(anomalies) if anomalies else ""
        success, err_msg = append_log(data, anomaly_message)
        if success:
            st.success("Log submitted successfully!")
        else:
            st.error(f"Error saving log: {err_msg}")
#!/usr/bin/env python3
import os
import datetime
import threading
import queue
import logging
import tkinter as tk
from tkinter import messagebox, ttk
from tkcalendar import DateEntry
import pandas as pd
from openpyxl import load_workbook
import win32com.client as win32

# ─── CONFIG ───────────────────────────────────────────────────────────────────
SRC_FILE = r"C:\path\to\your\CRIT_Master.xlsm"

consumer_dataProvider   = ["Alice Johnson", "Bob Smith"]
commercial_dataProvider = ["Carol Lee",    "Dan Patel"]
scheduleOwner           = ["Eve Zhang",    "Frank Müller"]
CR360Transformation     = ["Grace Kim",    "Hiro Tanaka"]

SHEET_MAP = {
    "Reg_Reporting_DP": [
        ("consumer_dataProvider",   consumer_dataProvider),
        ("commercial_dataProvider", commercial_dataProvider),
    ],
    "Reg_Reporting_SO": [
        ("scheduleOwner", scheduleOwner),
    ],
    "CR360": [
        ("CR360Transformation", CR360Transformation),
    ],
}

EMAIL_TO      = ["to1@example.com"]
EMAIL_CC      = ["cc1@example.com"]
EMAIL_SUBJECT = "CRIT Automated Summary"
# ─── END CONFIG ───────────────────────────────────────────────────────────────

# ─── logging setup ────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
logger = logging.getLogger(__name__)

def send_via_outlook(to, cc, subject, html_body, attachments):
    logger.info("Preparing Outlook email...")
    ol = win32.Dispatch('Outlook.Application')
    mail = ol.CreateItem(0)
    mail.To      = ";".join(to)
    mail.CC      = ";".join(cc)
    mail.Subject = subject
    mail.HTMLBody = html_body
    for f in attachments:
        logger.info(f"Attaching file {f}")
        mail.Attachments.Add(f)
    mail.Send()
    logger.info("Email sent via Outlook.")

def worker(start, end, q):
    try:
        logger.info("Worker thread starting.")
        # 1) build task list
        tasks = []
        for sheet, lists in SHEET_MAP.items():
            for list_name, emp_list in lists:
                tasks.append((sheet, list_name, emp_list))
        total_tasks = len(tasks)
        q.put(('init', total_tasks))

        now       = datetime.datetime.now()
        ts        = now.strftime("%Y%m%d_%H%M%S")
        downloads = os.path.join(os.path.expanduser("~"), "Downloads")

        attachments = []
        missing_info = {}

        # 2) preload DataFrames
        logger.info("Loading sheets into memory...")
        sheet_dfs = {
            sheet: pd.read_excel(
                SRC_FILE, sheet_name=sheet,
                engine="openpyxl", header=0, parse_dates=[0]
            )
            for sheet in SHEET_MAP
        }
        logger.info("Sheets loaded.")

        tasks_done = 0
        for sheet, list_name, emp_list in tasks:
            desc = f"{sheet} → {list_name}"
            logger.info(f"Starting task: {desc}")
            q.put(('task', desc))

            df_full = sheet_dfs[sheet]
            date_col = df_full.columns[0]
            emp_col  = df_full.columns[4]

            # 3) **Read-only preview** for row count & scan
            wb_prev = load_workbook(SRC_FILE, read_only=True, data_only=True)
            ws_prev = wb_prev[sheet]
            total_rows = ws_prev.max_row - 1
            q.put(('row_init', total_rows))

            to_delete = []
            rows_scanned = 0
            for r in range(2, total_rows + 2):
                rows_scanned += 1
                # **real-time update** every row
                q.put(('row_progress', rows_scanned))

                cell = ws_prev.cell(r, 1).value
                try:
                    dt = pd.to_datetime(cell).date()
                except:
                    dt = None
                emp = ws_prev.cell(r, 5).value
                if dt is None or dt < start or dt > end or emp not in emp_list:
                    to_delete.append(r)
            wb_prev.close()
            logger.info(f"{len(to_delete)} rows marked for deletion in {desc}.")

            # 4) figure out missing employees
            mask = (
                (df_full[date_col].dt.date >= start) &
                (df_full[date_col].dt.date <= end) &
                (df_full[emp_col].isin(emp_list))
            )
            present = set(df_full.loc[mask, emp_col].dropna().unique())
            missing = sorted(set(emp_list) - present)
            info = []
            for emp in missing:
                q.put(('employee', emp))
                df_emp = df_full[df_full[emp_col] == emp]
                if df_emp.empty:
                    info.append({"employee": emp, "last_date":"N/A", "last_row":"No record"})
                else:
                    last_dt = df_emp[date_col].max().date()
                    idx0    = df_emp[df_emp[date_col].dt.date == last_dt].index[0]
                    info.append({
                        "employee": emp,
                        "last_date": last_dt.strftime("%m/%d/%Y"),
                        "last_row":  idx0 + 2
                    })
            missing_info[list_name] = info
            logger.info(f"{len(missing)} missing for {list_name}.")

            # 5) delete rows & save filtered .xlsm
            wb = load_workbook(SRC_FILE, keep_vba=True)
            for s in list(wb.sheetnames):
                if s != sheet:
                    wb.remove(wb[s])
            ws = wb[sheet]
            for r in reversed(to_delete):
                ws.delete_rows(r)
            out_name = f"{list_name}_{ts}.xlsm"
            out_path = os.path.join(downloads, out_name)
            wb.save(out_path)
            wb.close()
            logger.info(f"Saved filtered workbook: {out_path}")
            attachments.append(out_path)

            tasks_done += 1
            q.put(('progress', tasks_done))

        # 6) build email HTML
        logger.info("Composing HTML email body.")
        html = ['<html><body><h1>Missing Entries</h1>']
        for ln, rows in missing_info.items():
            html.append(f"<h2>{ln}</h2>")
            if not rows:
                html.append("<p>All employees reported.</p>")
            else:
                html.append(
                    "<table border='1' cellpadding='4'>"
                    "<tr><th>Employee</th><th>Last Update</th><th>Last Row</th></tr>"
                )
                for r in rows:
                    html.append(
                        f"<tr><td>{r['employee']}</td>"
                        f"<td>{r['last_date']}</td>"
                        f"<td>{r['last_row']}</td></tr>"
                    )
                html.append("</table>")
        html.append("</body></html>")
        body = "\n".join(html)

        # 7) send via Outlook
        logger.info("Sending email via Outlook.")
        send_via_outlook(EMAIL_TO, EMAIL_CC, EMAIL_SUBJECT, body, attachments)

        q.put(('done', attachments, missing_info))
        logger.info("Worker thread completed successfully.")

    except Exception as e:
        logger.exception("Error in worker thread")
        q.put(('error', str(e)))

def start_process():
    start = start_cal.get_date()
    end   = end_cal.get_date()
    if start > end:
        messagebox.showerror("Error", "Start must be on or before End.")
        return

    root.start_time = datetime.datetime.now()
    elapsed_var.set("Elapsed: 00:00:00")

    btn_submit.config(state='disabled')
    progress_var.set("Initializing…")
    current_task_var.set("")
    row_progress_var.set("")
    current_emp_var.set("")
    prog_bar['value'] = 0
    row_bar['value'] = 0

    threading.Thread(target=worker, args=(start, end, q), daemon=True).start()
    root.after(100, poll_queue)
    update_elapsed()

def update_elapsed():
    if btn_submit['state'] == 'disabled':
        elapsed = datetime.datetime.now() - root.start_time
        elapsed_var.set(f"Elapsed: {str(elapsed).split('.')[0]}")
        root.after(1000, update_elapsed)

def poll_queue():
    try:
        msg = q.get_nowait()
    except queue.Empty:
        root.after(100, poll_queue)
        return

    kind = msg[0]
    if kind == 'init':
        total = msg[1]
        prog_bar['maximum'] = total
        progress_var.set(f"0 of {total} tasks")

    elif kind == 'task':
        current_task_var.set(f"Task: {msg[1]}")
        row_progress_var.set("")
        row_bar['value'] = 0

    elif kind == 'row_init':
        total_rows = msg[1]
        row_bar['maximum'] = total_rows
        row_progress_var.set(f"Rows: 0 of {total_rows}")

    elif kind == 'row_progress':
        done_rows = msg[1]
        row_bar['value'] = done_rows
        row_progress_var.set(f"Rows: {done_rows} of {row_bar['maximum']}")

    elif kind == 'employee':
        current_emp_var.set(f"Employee: {msg[1]}")

    elif kind == 'progress':
        done = msg[1]
        prog_bar['value'] = done
        progress_var.set(f"{done} of {prog_bar['maximum']} tasks")

    elif kind == 'done':
        btn_submit.config(state='normal')
        progress_var.set("All tasks done")
        current_task_var.set("")
        row_progress_var.set("")
        current_emp_var.set("")
        messagebox.showinfo("Finished", "Reports saved & email sent via Outlook.")

    elif kind == 'error':
        btn_submit.config(state='normal')
        progress_var.set("Error")
        current_task_var.set("")
        messagebox.showerror("Error in processing", msg[1])

    root.after(100, poll_queue)

# ─── BUILD UI ─────────────────────────────────────────────────────────────────
root = tk.Tk()
root.title("CRIT Filter & Emailer")
root.geometry("480x440")
root.resizable(False, False)

q = queue.Queue()
padx = 10
wrap = 460

tk.Label(root, text="Source (fixed):").pack(anchor="w", padx=padx, pady=(10,0))
tk.Label(root, text=SRC_FILE, wraplength=wrap, fg="gray").pack(anchor="w", padx=padx)

tk.Label(root, text="Start Date:").pack(anchor="w", padx=padx, pady=(10,0))
start_cal = DateEntry(root, date_pattern="mm/dd/yyyy"); start_cal.pack(padx=padx)

tk.Label(root, text="End Date:").pack(anchor="w", padx=padx, pady=(10,0))
end_cal = DateEntry(root, date_pattern="mm/dd/yyyy"); end_cal.pack(padx=padx)

btn_submit = tk.Button(
    root, text="Submit", command=start_process,
    bg="#4CAF50", fg="white",
    font=("Segoe UI", 12, "bold")
)
btn_submit.pack(pady=20, ipadx=10, ipady=5)

# Task progress
progress_var = tk.StringVar(value="Idle")
tk.Label(root, textvariable=progress_var, wraplength=wrap).pack(pady=(5,0))
prog_bar = ttk.Progressbar(root, orient="horizontal", length=450, mode="determinate")
prog_bar.pack(pady=(2,10))

# Current task
current_task_var = tk.StringVar(value="")
tk.Label(root, textvariable=current_task_var, fg="blue", wraplength=wrap).pack()

# Row progress
row_progress_var = tk.StringVar(value="")
tk.Label(root, textvariable=row_progress_var, wraplength=wrap).pack(pady=(5,0))
row_bar = ttk.Progressbar(root, orient="horizontal", length=450, mode="determinate")
row_bar.pack(pady=(2,10))

# Current employee
current_emp_var = tk.StringVar(value="")
tk.Label(root, textvariable=current_emp_var, wraplength=wrap).pack()

# Elapsed time
elapsed_var = tk.StringVar(value="Elapsed: 00:00:00")
tk.Label(root, textvariable=elapsed_var).pack(pady=(10,0))

root.mainloop()